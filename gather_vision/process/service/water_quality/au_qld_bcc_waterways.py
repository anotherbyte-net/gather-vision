import io
from datetime import datetime
from zoneinfo import ZoneInfo

from lxml import html
from openpyxl import load_workbook

from gather_vision import models as app_models
from gather_vision.process.component.http_client import HttpClient
from gather_vision.process.component.logger import Logger
from gather_vision.process.component.normalise import Normalise
from gather_vision.process.item.water_quality_measure import WaterQualityMeasure


class AuQldBccWaterways:
    code = "au_qld_bcc_waterways"
    title = "Brisbane City Council waterways quality monitoring"
    short_title = "BCC waterways quality"
    page_url = "https://www.brisbane.qld.gov.au/clean-and-green/natural-environment-and-water/water/water-quality-monitoring"

    def __init__(
        self,
        logger: Logger,
        http_client: HttpClient,
        normalise: Normalise,
        tz: ZoneInfo,
    ):
        self._logger = logger
        self._http_client = http_client
        self._normalise = normalise
        self._tz = tz

        self._source = app_models.InformationSource.objects.get(name=self.code)

    def update_measurements(self):
        url, text = self._get_url()
        raw = self._gather_rows(url)
        measures = self._get_measures(raw)

        for measure in measures:
            yield measure

    def _get_measures(self, raw: dict[str, dict[str, list]]):
        for name, data in raw.items():
            headers = {}
            for item in data["headers"]:
                for (col, header) in item:
                    headers[col] = header

            site = {}
            for body in data["body"]:
                for col, value in body:
                    header = headers[col]
                    if header == "Site no.":
                        site["site_number"] = value
                    elif header == "Site name":
                        site["site_name"] = value
                    elif header == "Long":
                        site["location_longitude"] = value
                    elif header == "Lat":
                        site["location_latitude"] = value
                    elif header == "Location description":
                        site["location_description"] = value
                    elif isinstance(header, datetime):
                        measure_date = header.replace(tzinfo=self._tz)
                        yield WaterQualityMeasure.build_from_excel(
                            site, measure_date, value
                        )
                    else:
                        raise ValueError(
                            f"Unknown header '{header}' with value '{value}'."
                        )

    def _get_url(self):
        r = self._http_client.get(self.page_url)
        tree = html.fromstring(r.text)

        if tree is None:
            raise ValueError("No html available.")

        link = tree.xpath('//a[contains(@href, "xlsx")]')
        if len(link) != 1:
            raise ValueError("Found other than 1 link in html.")
        link = link[0]

        url = link.xpath("./@href")
        if len(url) != 1:
            raise ValueError("Found other than 1 url in html.")
        url = url[0]

        text = link.text

        return url, text

    def _gather_rows(self, url: str):
        r = self._http_client.get(url)
        filename = io.BytesIO(r.content)

        result = {}

        wb = load_workbook(filename=filename, data_only=True, read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            headers = []
            body = []

            for row in ws.rows:
                cells = [cell for cell in row if cell.value]
                cell_count = len(cells)

                if cell_count < 6:
                    continue

                cell_first = cells[0]
                if (
                    isinstance(cell_first.value, int)
                    or str(cell_first.value)[0].isnumeric()
                ):
                    body.append([(cell.column, cell.value) for cell in cells])
                else:
                    headers.append([(cell.column, cell.value) for cell in cells])

            result[sheet_name] = {"headers": headers, "body": body}

        # Close the workbook after reading
        wb.close()

        return result
