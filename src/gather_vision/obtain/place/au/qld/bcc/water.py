import dataclasses
from datetime import datetime, timezone
import io
import typing
from zoneinfo import ZoneInfo

from gather_vision.apps.water import models as water_models
from gather_vision.obtain.core import data
from openpyxl import load_workbook


@dataclasses.dataclass(frozen=True)
class BrisbaneCityCouncilWaterQualityItem(data.GatherDataItem):
    # TODO: attributes
    @classmethod
    def build(
        cls,
        site_info: dict,
        measure_date: datetime,
        measure_value: typing.Union[str, int],
    ):
        if isinstance(measure_value, int):
            observation_value = measure_value
            observation_status = water_models.Measure.QUALITY_VALID
        elif str(measure_value or "").strip() == "NT":
            observation_value = None
            observation_status = water_models.Measure.QUALITY_NOT_TESTED
        elif str(measure_value or "").strip().startswith(">") or str(
            measure_value or ""
        ).strip().startswith("<"):
            observation_value = int(
                str(measure_value or "").strip()[1:].replace(",", ""), base=10
            )
            observation_status = water_models.Measure.QUALITY_VALID
        else:
            observation_value = None
            observation_status = water_models.Measure.QUALITY_INVALID

        return BrisbaneCityCouncilWaterQualityItem(
            observation_date=measure_date,
            observation_value=observation_value,
            observation_status=observation_status,
            **site_info,
        )


@dataclasses.dataclass(frozen=True)
class BrisbaneCityCouncilWaterLevelItem(data.GatherDataItem):
    pass


class BrisbaneCityCouncilWaterWebData(data.WebData):
    _tz = ZoneInfo("Australia/Brisbane")

    # water quality
    water_url = "https://www.brisbane.qld.gov.au/clean-and-green/natural-environment-and-water/water/water-quality-monitoring"

    # water levels in dams
    @property
    def tags(self) -> dict[str, str]:
        return {
            "country": "Australia",
            "region": "Queensland",
            "district": "Brisbane City Council",
            "locality": "City of Brisbane",
            "data_source_location": "web",
            "data_source_category": "water",
        }

    @property
    def name(self):
        return "au-qld-bcc-water"

    def initial_urls(self) -> typing.Iterable[str]:
        return [self.water_url]

    def web_resources(
        self, web_data: data.WebDataAvailable
    ) -> typing.Iterable[typing.Union[data.GatherDataRequest, data.GatherDataItem]]:
        url = web_data.response_url
        if url.startswith(self.water_url):
            yield self._get_excel_file(web_data)

        elif ".xls" in url:
            raw = self._get_rows(web_data)
            for measure in self._get_measures(raw):
                yield measure

    def _get_excel_file(
        self, web_data: data.WebDataAvailable
    ) -> data.GatherDataRequest:
        # link text
        links = web_data.selector.xpath('//a[contains(@href, "xlsx")]')
        if len(links) != 1:
            raise ValueError("Found other than 1 link in html.")
        link_text = links[0].css("::text").get()

        # link url
        urls = links[0].xpath("./@href")
        if len(urls) != 1:
            raise ValueError("Found other than 1 url in html.")
        link_url = urls[0].get()

        return data.GatherDataRequest(
            url=link_url,
            data={"link_text": link_text},
        )

    def _get_rows(self, web_data: data.WebDataAvailable) -> dict:
        filename = io.BytesIO(web_data.body_raw)

        known_headers = ["Site no.", "Site name", "Long", "Lat", "Location description"]

        wb = None
        try:
            wb = load_workbook(filename=filename, data_only=True, read_only=True)

            result = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                headers = []
                body = []

                for row in ws.rows:
                    cells = [cell for cell in row if cell.value]
                    cell_count = len(cells)

                    if cell_count < 6:
                        any_dates = any(
                            [isinstance(cell.value, datetime) for cell in cells]
                        )
                        known_header = any(
                            [cell.value in known_headers for cell in cells]
                        )
                        if not any_dates and not known_header:
                            continue

                    cell_first = cells[0]
                    if (
                        isinstance(cell_first.value, int)
                        or str(cell_first.value)[0].isnumeric()
                    ):
                        body.append([(cell.column, cell.value) for cell in cells])
                    else:
                        headers.append([(cell.column, cell.value) for cell in cells])

                result[sheet_name] = {"headers": headers, "body": body}

        finally:
            # Close the workbook after reading
            if wb:
                wb.close()

        return result

    def _get_measures(
        self, raw: dict[str, dict[str, list]]
    ) -> typing.Iterable[BrisbaneCityCouncilWaterQualityItem]:
        for name, info in raw.items():
            headers = {}
            for item in info["headers"]:
                for col, header in item:
                    headers[col] = header

            site = {}
            for body in info["body"]:
                for col, value in body:
                    header = headers[col]
                    if header == "Site no.":
                        site["site_number"] = value
                    elif header == "Site name":
                        site["site_name"] = value
                    elif header == "Long":
                        site["location_longitude"] = value
                    elif header == "Lat":
                        site["location_latitude"] = value
                    elif header == "Location description":
                        site["location_description"] = value
                    elif isinstance(header, datetime):
                        measure_date = header.replace(tzinfo=self._tz)
                        yield None
                        # TODO: item build
                        # BrisbaneCityCouncilWaterQualityItem.build(
                        #     site, measure_date, value
                        # )
                    else:
                        raise ValueError(
                            f"Unknown header '{header}' with value '{value}'."
                        )
